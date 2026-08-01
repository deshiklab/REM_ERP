"""REM ERP Backend — Flask app: auth (session + RBAC), REST APIs mirroring the V10 prototype."""
import json, re, io, csv, uuid
from datetime import datetime, date
from functools import wraps
from flask import Flask, jsonify, request, session, Response
from werkzeug.security import check_password_hash
from db import get_db, rows_to_dicts, row_to_dict, log_activity

app = Flask(__name__)
app.secret_key = 'rem-erp-dev-secret-change-in-prod'

# ── CORS (prototype served from a different origin via tunnel) ────────
@app.after_request
def _cors(resp):
    resp.headers['Access-Control-Allow-Origin'] = request.headers.get('Origin', '*')
    resp.headers['Access-Control-Allow-Credentials'] = 'true'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    if request.method == 'OPTIONS':
        return resp
    return resp

# ── RBAC (mirrors prototype ROLE_MODULES) ─────────────────────────────
ROLE_MODULES = {
    'Super Admin': 'all',
    'Sales Agent': ['leads', 'customers', 'bookings', 'dashboard'],
    'Site Engineer': ['projects', 'assets', 'dashboard'],
    'Finance': ['invoices', 'payments', 'dues', 'customers', 'assets', 'license', 'dashboard'],
}

def _auth_user():
    """Resolve user id from session cookie OR Authorization: Bearer <token>."""
    if 'uid' in session:
        return session['uid']
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        tok = auth[7:].strip()
        conn = get_db()
        r = conn.execute("SELECT user_id FROM api_tokens WHERE token=?", (tok,)).fetchone()
        conn.close()
        if r:
            return r['user_id']
    return None

def require_login(f):
    @wraps(f)
    def wrapper(*a, **kw):
        if _auth_user() is None:
            return jsonify({'error': 'Unauthorized — login required'}), 401
        return f(*a, **kw)
    return wrapper

def require_module(module):
    def deco(f):
        @wraps(f)
        def wrapper(*a, **kw):
            uid = _auth_user()
            if uid is None:
                return jsonify({'error': 'Unauthorized'}), 401
            role = session.get('role')
            if role is None:
                conn = get_db()
                u = conn.execute("SELECT role FROM users WHERE id=?", (uid,)).fetchone()
                conn.close()
                role = u['role'] if u else None
            allowed = ROLE_MODULES.get(role, [])
            if allowed != 'all' and module not in allowed:
                return jsonify({'error': f'Access denied — {module} not available for {role}'}), 403
            return f(*a, **kw)
        return wrapper
    return deco

def _user(uid=None):
    uid = uid or _auth_user()
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE id=?", (uid,)).fetchone()
    conn.close()
    return row_to_dict(u)

def _audit(action, module, entity, entity_id, details=''):
    u = session.get('name', 'System')
    log_activity(u, action, module, entity, entity_id, details)

# ── AUTH ──────────────────────────────────────────────────────────────
@app.post('/api/login')
def login():
    data = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    u = conn.execute("SELECT * FROM users WHERE email=?", (((data or {}).get('email') or '').strip(),)).fetchone()
    conn.close()
    if not u or not check_password_hash(u['password_hash'], data.get('password', '')):
        return jsonify({'error': 'Invalid email or password'}), 401
    session.clear()
    session['uid'] = u['id']; session['name'] = u['name']; session['role'] = u['role']
    conn = get_db()
    conn.execute("UPDATE users SET last_login=? WHERE id=?", (datetime.utcnow().isoformat(), u['id']))
    tok = uuid.uuid4().hex
    conn.execute("INSERT INTO api_tokens(token,user_id,created_at) VALUES(?,?,?)",
                 (tok, u['id'], datetime.utcnow().isoformat()))
    conn.commit(); conn.close()
    log_activity(u['name'], 'Login', 'System', 'auth', str(u['id']))
    return jsonify({'ok': True, 'token': tok, 'user': {'id': u['id'], 'name': u['name'], 'email': u['email'], 'role': u['role']}})

@app.get('/api/me')
@require_login
def me():
    return jsonify({'user': _user()})

# ── DASHBOARD ─────────────────────────────────────────────────────────
@app.get('/api/dashboard')
@require_login
def dashboard():
    conn = get_db()
    def one(q, *a): return conn.execute(q, a).fetchone()[0] or 0
    stats = {
        'projects': one("SELECT COUNT(*) FROM projects"),
        'portfolio': one("SELECT COALESCE(SUM(budget),0) FROM projects"),
        'leads': one("SELECT COUNT(*) FROM leads"),
        'bookings': one("SELECT COUNT(*) FROM bookings"),
        'sales': one("SELECT COALESCE(SUM(price),0) FROM bookings"),
        'invoices': one("SELECT COUNT(*) FROM invoices"),
        'invoices_unpaid': one("SELECT COUNT(*) FROM invoices WHERE status NOT IN ('Paid')"),
        'payments': one("SELECT COUNT(*) FROM payments"),
        'payments_cleared': one("SELECT COALESCE(SUM(amount),0) FROM payments WHERE status='Cleared'"),
        'dues_outstanding': one("SELECT COALESCE(SUM(due),0) FROM dues"),
        'cash_in': one("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='Inflow' AND status IN ('Received','Cleared')"),
        'cash_out': one("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='Outflow'"),
        'assets_nbv': one("SELECT COALESCE(SUM(cost-accum_dep),0) FROM fixed_assets"),
        'employees': one("SELECT COUNT(*) FROM users"),
    }
    conn.close()
    return jsonify({'stats': stats})

# ── GENERIC CRUD FACTORY ──────────────────────────────────────────────
def _crud(table, module, fields, id_field='id', pk=None):
    def _list():
        conn = get_db()
        q = f"SELECT * FROM {table}"
        args = []
        status = request.args.get('status')
        if status:
            q += f" WHERE status=?"; args.append(status)
        q += " ORDER BY id DESC" if id_field == 'id' else " ORDER BY rowid DESC"
        rows = conn.execute(q, args).fetchall(); conn.close()
        return jsonify(rows_to_dicts(rows))
    def _create():
        data = request.get_json(force=True, silent=True) or {}
        keys = [k for k in fields if k in data]
        vals = [data[k] for k in keys]
        conn = get_db()
        cur = conn.execute(f"INSERT INTO {table}({','.join(keys)}) VALUES({','.join(['?']*len(keys))})", vals)
        conn.commit()
        new_id = cur.lastrowid
        if pk is not None:
            new_id = data.get(pk, new_id)
        conn.close()
        _audit('Created', module, table, str(new_id), json.dumps(data)[:200])
        return jsonify({'ok': True, 'id': str(new_id)}), 201
    def _update(rid):
        data = request.get_json(force=True, silent=True) or {}
        sets = [f"{k}=?" for k in data if k in fields]
        if not sets: return jsonify({'error': 'No valid fields'}), 400
        conn = get_db()
        conn.execute(f"UPDATE {table} SET {','.join(sets)} WHERE {id_field}=?", list(data.values()) + [rid])
        conn.commit(); conn.close()
        _audit('Updated', module, table, str(rid), json.dumps(data)[:200])
        return jsonify({'ok': True})
    def _delete(rid):
        conn = get_db()
        conn.execute(f"DELETE FROM {table} WHERE {id_field}=?", (rid,))
        conn.commit(); conn.close()
        _audit('Deleted', module, table, str(rid))
        return jsonify({'ok': True})
    def _get(rid):
        conn = get_db()
        r = conn.execute(f"SELECT * FROM {table} WHERE {id_field}=?", (rid,)).fetchone()
        conn.close()
        return jsonify(row_to_dict(r) or {'error': 'not found'}), (200 if r else 404)
    return _list, _get, _create, _update, _delete

def _register(table, module, fields, id_field='id', pk=None, perms=('*',)):
    lst, get, cre, upd, dele = _crud(table, module, fields, id_field, pk)
    app.add_url_rule(f'/api/{module}', 'list_' + module, require_module(module)(lst), methods=['GET'])
    app.add_url_rule(f'/api/{module}/<string:rid>', 'get_' + module, require_module(module)(get), methods=['GET'])
    app.add_url_rule(f'/api/{module}', 'create_' + module, require_module(module)(cre), methods=['POST'])
    app.add_url_rule(f'/api/{module}/<string:rid>', 'update_' + module, require_module(module)(upd), methods=['PUT'])
    app.add_url_rule(f'/api/{module}/<string:rid>', 'delete_' + module, require_module(module)(dele), methods=['DELETE'])

_register('leads', 'leads', ['name','phone','email','property','status','priority','type','source','value','owner','next_follow_up','created_at'])
_register('customers', 'customers', ['name','phone','email','property','type','status','dues_num','project'])
_register('projects', 'projects', ['code','name','location','status','progress','budget','manager','type','plots','units'])
_register('dues', 'dues', ['customer','project','unit','total_price','paid','due','due_date','status','bucket','days_overdue','phone'])
_register('fixed_assets', 'assets', ['code','name','category','purchase_date','cost','salvage','useful_life','accum_dep','location','status'], pk='id')

# ── BOOKINGS (string PK) ──────────────────────────────────────────────
@app.get('/api/bookings')
@require_module('bookings')
def list_bookings():
    conn = get_db(); rows = conn.execute("SELECT * FROM bookings ORDER BY date DESC").fetchall(); conn.close()
    return jsonify(rows_to_dicts(rows))

@app.post('/api/bookings')
@require_module('bookings')
def create_booking():
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    mx = conn.execute("SELECT MAX(CAST(REPLACE(id,'BKG-','') AS INTEGER)) FROM bookings").fetchone()[0] or 100
    nid = f"BKG-{int(mx)+1}"
    conn.execute("INSERT INTO bookings(id,client,property,unit,price,advance,status,type,terms,sched_start,date) VALUES(?,?,?,?,?,?,?,?,?,?,?)",
                 (nid, d.get('client',''), d.get('property',''), d.get('unit',''), int(d.get('price') or 0),
                  int(d.get('advance') or 0), d.get('status','Pending Review'), d.get('type','Flat'),
                  d.get('terms',''), d.get('sched_start', datetime.utcnow().date().isoformat()),
                  datetime.utcnow().strftime('%b %d, %Y')))
    conn.commit(); conn.close()
    _audit('Created', 'bookings', 'bookings', nid)
    return jsonify({'ok': True, 'id': nid}), 201

@app.put('/api/bookings/<rid>')
@require_module('bookings')
def update_booking(rid):
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    conn.execute("UPDATE bookings SET client=?,property=?,unit=?,price=?,advance=?,status=?,type=?,terms=?,sched_start=? WHERE id=?",
                 (d.get('client',''), d.get('property',''), d.get('unit',''), int(d.get('price') or 0),
                  int(d.get('advance') or 0), d.get('status',''), d.get('type',''), d.get('terms',''),
                  d.get('sched_start',''), rid))
    conn.commit(); conn.close(); _audit('Updated', 'bookings', 'bookings', rid)
    return jsonify({'ok': True})

@app.delete('/api/bookings/<rid>')
@require_module('bookings')
def delete_booking(rid):
    conn = get_db(); conn.execute("DELETE FROM bookings WHERE id=?", (rid,)); conn.commit(); conn.close()
    _audit('Deleted', 'bookings', 'bookings', rid)
    return jsonify({'ok': True})

# ── INVOICES (VAT net computation mirror) ─────────────────────────────
def _inv_net(amount, vat_r, tds_r, ait_r):
    vat = round(amount * vat_r / 100); tds = round(amount * tds_r / 100); ait = round(amount * ait_r / 100)
    return vat, tds, ait, max(0, amount + vat - tds - ait)

@app.get('/api/invoices')
@require_module('invoices')
def list_invoices():
    conn = get_db(); rows = conn.execute("SELECT * FROM invoices ORDER BY issued_date DESC").fetchall(); conn.close()
    return jsonify(rows_to_dicts(rows))

@app.post('/api/invoices')
@require_module('invoices')
def create_invoice():
    d = request.get_json(force=True, silent=True) or {}
    amount = int(d.get('amount') or 0)
    vat_r = int(d.get('vat_rate') or 0); tds_r = int(d.get('tds_rate') or 0); ait_r = int(d.get('ait_rate') or 0)
    vat, tds, ait, net = _inv_net(amount, vat_r, tds_r, ait_r)
    conn = get_db()
    yr = datetime.utcnow().year
    mx = conn.execute("SELECT MAX(CAST(SUBSTR(id,10) AS INTEGER)) FROM invoices WHERE id LIKE ?", (f"INV-{yr}-%",)).fetchone()[0] or 0
    nid = f"INV-{yr}-{int(mx)+1:04d}"
    conn.execute("INSERT INTO invoices(id,client,project,unit,amount,vat_rate,tds_rate,ait_rate,vat,tds,ait,net,challan,status,due_date,issued_date,desc,type) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                 (nid, d.get('client',''), d.get('project',''), d.get('unit',''), amount, vat_r, tds_r, ait_r,
                  vat, tds, ait, net, d.get('challan',''), d.get('status','Draft'), d.get('due_date',''),
                  d.get('issued_date', datetime.utcnow().date().isoformat()), d.get('desc',''), 'Sales'))
    conn.commit(); conn.close()
    _audit('Created', 'invoices', 'invoices', nid, f"net={net}")
    return jsonify({'ok': True, 'id': nid, 'vat': vat, 'tds': tds, 'ait': ait, 'net': net}), 201

# ── PAYMENTS (ripple mirror: invoice status + dues + txn) ─────────────
def _recompute_invoice(conn, inv_id):
    row = conn.execute("SELECT * FROM invoices WHERE id=?", (inv_id,)).fetchone()
    if not row: return
    paid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE invoice_id=? AND status='Cleared'", (inv_id,)).fetchone()[0]
    net = row['net'] or row['amount']
    status = 'Paid' if paid >= net else ('Partial' if paid > 0 else ('Overdue' if row['status'] == 'Overdue' else ('Draft' if row['status'] == 'Draft' else 'Sent')))
    conn.execute("UPDATE invoices SET status=? WHERE id=?", (status, inv_id))

@app.post('/api/payments')
@require_module('payments')
def create_payment():
    d = request.get_json(force=True, silent=True) or {}
    amount = int(d.get('amount') or 0)
    conn = get_db()
    mx = conn.execute("SELECT MAX(CAST(REPLACE(id,'PAY-','') AS INTEGER)) FROM payments").fetchone()[0] or 0
    pid = f"PAY-{int(mx)+1:03d}"
    inv_id = d.get('invoice_id') or None
    conn.execute("INSERT INTO payments(id,invoice_id,client,amount,date,method,reference,status,notes) VALUES(?,?,?,?,?,?,?,?,?)",
                 (pid, inv_id, d.get('client',''), amount, d.get('date', datetime.utcnow().date().isoformat()),
                  d.get('method','Bank Transfer'), d.get('reference',''), d.get('status','Cleared'), d.get('notes','')))
    if inv_id and amount > 0:
        _recompute_invoice(conn, inv_id)
        # dues ripple
        client = d.get('client','')
        if client:
            due = conn.execute("SELECT * FROM dues WHERE customer=?", (client,)).fetchone()
            if due:
                new_due = max(0, due['due'] - amount)
                status = 'Paid' if new_due <= 0 else 'Upcoming'
                bucket = 'Cleared' if new_due <= 0 else 'On Track'
                conn.execute("UPDATE dues SET paid=paid+?, due=?, status=?, bucket=?, days_overdue=0 WHERE id=?",
                             (amount, new_due, status, bucket, due['id']))
        # cash-flow transaction
        conn.execute("INSERT INTO transactions(id,date,desc,client,project,type,category,status,amount) VALUES(?,?,?,?,?,?,?,?,?)",
                     (f"RCP-{1000+int(mx)+1}", d.get('date', datetime.utcnow().date().isoformat()),
                      f"Payment - {d.get('client','')}", d.get('client',''), '', 'Inflow', 'Payment Received', 'Received', amount))
    conn.commit(); conn.close()
    _audit('Created', 'payments', 'payments', pid, f"amount={amount}")
    return jsonify({'ok': True, 'id': pid}), 201

@app.put('/api/payments/<rid>')
@require_module('payments')
def update_payment(rid):
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db()
    conn.execute("UPDATE payments SET invoice_id=?,client=?,amount=?,date=?,method=?,reference=?,status=?,notes=? WHERE id=?",
                 (d.get('invoice_id'), d.get('client',''), int(d.get('amount') or 0), d.get('date',''),
                  d.get('method','Bank Transfer'), d.get('reference',''), d.get('status',''), d.get('notes',''), rid))
    if d.get('invoice_id'):
        _recompute_invoice(conn, d['invoice_id'])
    conn.commit(); conn.close(); _audit('Updated', 'payments', 'payments', rid)
    return jsonify({'ok': True})

# ── LICENSE ───────────────────────────────────────────────────────────
@app.get('/api/license')
@require_module('license')
def get_license():
    conn = get_db(); r = conn.execute("SELECT * FROM license WHERE id=1").fetchone(); conn.close()
    lic = row_to_dict(r)
    lic['installments'] = json.loads(lic.get('installments') or '[]')
    lic['checklist'] = json.loads(lic.get('checklist') or '[]')
    lic['paid'] = sum(i['amount'] for i in lic['installments'] if i['status'] == 'Paid')
    lic['pct'] = round(lic['paid'] / lic['contract'] * 100) if lic['contract'] else 0
    return jsonify(lic)

@app.post('/api/license/status')
@require_module('license')
def set_license_status():
    d = request.get_json(force=True, silent=True) or {}
    conn = get_db(); conn.execute("UPDATE license SET status=? WHERE id=1", (d.get('status', 'Active'),)); conn.commit(); conn.close()
    _audit('Updated', 'license', 'license', '1', d.get('status', ''))
    return jsonify({'ok': True})

# ── REPORTS / EXPORTS ─────────────────────────────────────────────────
@app.get('/api/reports/unpaid-invoices')
@require_module('invoices')
def unpaid_invoices():
    conn = get_db()
    rows = conn.execute("SELECT * FROM invoices WHERE status NOT IN ('Paid') ORDER BY due_date").fetchall()
    out = []
    for r in rows:
        paid = conn.execute("SELECT COALESCE(SUM(amount),0) FROM payments WHERE invoice_id=? AND status='Cleared'", (r['id'],)).fetchone()[0]
        out.append({**dict(r), 'paid': paid, 'outstanding': max(0, (r['net'] or r['amount']) - paid)})
    conn.close()
    return jsonify(out)

# ── PHASE 3: DOC STORE SYNC (prototype <-> server) ────────────────────
@app.get('/api/bootstrap')
@require_login
def bootstrap():
    uid = _auth_user()
    conn = get_db()
    rows = conn.execute("SELECT collection, data FROM doc_store ORDER BY collection").fetchall()
    collections = {}
    for r in rows:
        try:
            collections.setdefault(r['collection'], []).append(json.loads(r['data']))
        except Exception:
            pass
    # scalar (object) collections come back as arrays too; client merges by id
    meta = {'server_time': datetime.utcnow().isoformat() + 'Z',
            'collections': sorted(collections.keys()),
            'counts': {k: len(v) for k, v in collections.items()}}
    u = _user(uid)
    return jsonify({'ok': True, 'meta': meta, 'collections': collections, 'user': u})

@app.post('/api/sync')
@require_login
def sync():
    data = request.get_json(force=True, silent=True) or {}
    cols = data.get('collections') or {}
    conn = get_db()
    now = datetime.utcnow().isoformat() + 'Z'
    n = 0
    for name, rows in cols.items():
        if isinstance(rows, dict):
            rows = [rows]
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            rid = str(row.get('id') or row.get('code') or row.get('uid') or row.get('key') or f"r{n}")
            conn.execute(
                "INSERT INTO doc_store(collection,id,data,updated_at) VALUES(?,?,?,?) "
                "ON CONFLICT(collection,id) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at",
                (name, rid, json.dumps(row, ensure_ascii=False), now))
            n += 1
    conn.commit()
    last = {'collection': '_meta', 'id': '_last_sync', 'data': json.dumps({'at': now, 'rows': n}),
            'updated_at': now}
    conn.execute("INSERT INTO doc_store(collection,id,data,updated_at) VALUES(?,?,?,?) "
                 "ON CONFLICT(collection,id) DO UPDATE SET data=excluded.data, updated_at=excluded.updated_at",
                 ('_meta', '_last_sync', last['data'], now))
    conn.commit(); conn.close()
    _audit('Sync', 'System', 'doc_store', str(n) + ' rows')
    return jsonify({'ok': True, 'rows': n})

@app.post('/api/logout')
def logout():
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Bearer '):
        conn = get_db()
        conn.execute("DELETE FROM api_tokens WHERE token=?", (auth[7:].strip(),))
        conn.commit(); conn.close()
    session.clear()
    return jsonify({'ok': True})

# ── PHASE 4: REPORTS (PDF / CSV / Excel) ──────────────────────────────
def _doc_collection(name):
    conn = get_db()
    rows = conn.execute("SELECT data FROM doc_store WHERE collection=?", (name,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        try:
            d = json.loads(r['data'])
            if isinstance(d, dict):
                out.append(d)
        except Exception:
            pass
    return out

def _money(n):
    try:
        return f"Tk {int(n):,}"
    except Exception:
        return "Tk 0"

def _doc_invoice(rid):
    invs = _doc_collection('invoices')
    for i in invs:
        if str(i.get('id')) == rid:
            return i
    return None

@app.get('/api/reports/invoice/<rid>.pdf')
@require_login
def invoice_pdf(rid):
    inv = _doc_invoice(rid)
    if not inv:
        return jsonify({'error': 'Invoice not found in sync store'}), 404
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except Exception:
        return jsonify({'error': 'reportlab not installed on server'}), 500
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=36, rightMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Title'], fontSize=16, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey, spaceAfter=10)
    small = ParagraphStyle('small', parent=styles['Normal'], fontSize=9)
    amount = int(inv.get('amount') or 0)
    vat_r = int(inv.get('vat_rate') or 0); tds_r = int(inv.get('tds_rate') or 0); ait_r = int(inv.get('ait_rate') or 0)
    vat = int(inv.get('vat') or 0); tds = int(inv.get('tds') or 0); ait = int(inv.get('ait') or 0)
    net = int(inv.get('net') or 0)
    if not vat and vat_r: vat = round(amount * vat_r / 100)
    if not tds and tds_r: tds = round(amount * tds_r / 100)
    if not ait and ait_r: ait = round(amount * ait_r / 100)
    if not net: net = max(0, amount + vat - tds - ait)
    story = []
    story.append(Paragraph('REM ERP — Tax Invoice', h1))
    story.append(Paragraph(f"Invoice {inv.get('id','')}  ·  Status: {inv.get('status','Draft')}  ·  Issued: {inv.get('issued_date','')}  ·  Due: {inv.get('due_date','')}", sub))
    story.append(Paragraph(f"<b>Client:</b> {inv.get('client','')}", small))
    story.append(Paragraph(f"<b>Project:</b> {inv.get('project','')}  ·  <b>Unit:</b> {inv.get('unit','')}", small))
    if inv.get('challan'):
        story.append(Paragraph(f"<b>VAT Challan Ref:</b> {inv.get('challan','')}", small))
    story.append(Spacer(1, 14))
    tdata = [
        ['Description', 'Amount (Tk)'],
        [inv.get('desc') or f'Sales — {inv.get("client","")}', f"{amount:,.0f}"],
        [f'VAT @ {vat_r}%', f"{vat:,.0f}"],
        [f'TDS @ {tds_r}%', f"-{tds:,.0f}"],
        [f'AIT @ {ait_r}%', f"-{ait:,.0f}"],
        ['Net Payable', f"{net:,.0f}"],
    ]
    tbl = Table(tdata, colWidths=[330, 110])
    tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F80ED')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#eef3ff')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#d0d8e8')),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 12))
    story.append(Paragraph('Payment terms: as per booking agreement. This is a computer-generated invoice.', sub))
    doc.build(story)
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': f'attachment; filename={rid}.pdf'})

@app.get('/api/reports/vat-register.pdf')
@require_login
def vat_register_pdf():
    invs = _doc_collection('invoices')
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except Exception:
        return jsonify({'error': 'reportlab not installed on server'}), 500
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=30, rightMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('h1', parent=styles['Title'], fontSize=14, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=8, textColor=colors.grey, spaceAfter=8)
    rows = [['Invoice', 'Client', 'Subtotal', 'VAT%', 'VAT', 'TDS', 'AIT', 'Net', 'Challan', 'Status']]
    tv = tt = ta = tn = 0
    for i in invs:
        amount = int(i.get('amount') or 0); vat = int(i.get('vat') or 0)
        tds = int(i.get('tds') or 0); ait = int(i.get('ait') or 0)
        net = int(i.get('net') or 0)
        if not net: net = max(0, amount + vat - tds - ait)
        rows.append([str(i.get('id','')), str(i.get('client',''))[:22], f"{amount:,.0f}",
                     str(i.get('vat_rate') or 0), f"{vat:,.0f}", f"{tds:,.0f}", f"{ait:,.0f}",
                     f"{net:,.0f}", str(i.get('challan','')), str(i.get('status',''))])
        tv += vat; tt += tds; ta += ait; tn += net
    rows.append(['TOTAL', f"{len(invs)} invoices", '', '', f"{tv:,.0f}", f"{tt:,.0f}", f"{ta:,.0f}", f"{tn:,.0f}", '', ''])
    tbl = Table(rows, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F80ED')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#eef3ff')),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#d0d8e8')),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
    ]))
    story = [Paragraph('REM ERP — VAT Register', h1),
             Paragraph(f'{len(invs)} invoices · generated {datetime.utcnow().strftime("%d %b %Y %H:%M")} UTC · Net VAT {_money(tv)}', sub),
             tbl]
    doc.build(story)
    return Response(buf.getvalue(), mimetype='application/pdf',
                    headers={'Content-Disposition': 'attachment; filename=vat-register.pdf'})

@app.get('/api/reports/csv/<collection>.csv')
@require_login
def collection_csv(collection):
    rows = _doc_collection(collection)
    headers = []
    for r in rows:
        for k in r.keys():
            if k not in headers:
                headers.append(k)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(headers)
    for r in rows:
        w.writerow([r.get(h, '') if not isinstance(r.get(h), (dict, list)) else json.dumps(r.get(h), ensure_ascii=False) for h in headers])
    return Response('\ufeff' + buf.getvalue(), mimetype='text/csv; charset=utf-8',
                    headers={'Content-Disposition': f'attachment; filename={collection}.csv'})

@app.get('/api/reports/xlsx/<collection>.xlsx')
@require_login
def collection_xlsx(collection):
    rows = _doc_collection(collection)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        return jsonify({'error': 'openpyxl not installed on server'}), 500
    headers = []
    for r in rows:
        for k in r.keys():
            if k not in headers:
                headers.append(k)
    wb = Workbook()
    ws = wb.active
    ws.title = collection[:28] or 'data'
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='2F80ED')
    for r in rows:
        ws.append([r.get(h, '') if not isinstance(r.get(h), (dict, list)) else json.dumps(r.get(h), ensure_ascii=False) for h in headers])
    buf = io.BytesIO()
    wb.save(buf)
    return Response(buf.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename={collection}.xlsx'})

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=False)
